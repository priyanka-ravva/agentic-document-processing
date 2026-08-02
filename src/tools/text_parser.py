"""Text and structured-file extraction tool."""

import csv
import json
from pathlib import Path



def extract_text_from_text_document(file_path: str) -> str:
    """Extract textual content from supported text and structured document formats."""

    path = Path(file_path)
    extension = path.suffix.lower()

    if extension == ".txt":
        return path.read_text(encoding="utf-8", errors="replace").strip()

    if extension == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        return json.dumps(payload, indent=2, ensure_ascii=False)

    if extension == ".csv":
        return _read_csv_as_text(path)

    if extension == ".xlsx":
        return _read_xlsx_as_text(path)

    if extension == ".docx":
        return _read_docx_as_text(path)

    raise ValueError(
        f"Text parser only supports .txt, .json, .csv, .xlsx, and .docx files. Got: {extension}"
    )


def _read_csv_as_text(path: Path) -> str:
    """Serialize CSV rows into readable text."""

    rows: list[str] = []
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            rows.append(" | ".join(str(cell) for cell in row))
    return "\n".join(rows).strip()


def _read_xlsx_as_text(path: Path) -> str:
    """Serialize XLSX workbook sheets into readable text."""

    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required to parse .xlsx files. Install it with `pip install openpyxl`."
        ) from exc

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_texts: list[str] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        row_texts: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if cell is None else str(cell) for cell in row]
            if any(cell.strip() for cell in cells):
                row_texts.append(" | ".join(cells))
        if row_texts:
            sheet_texts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(row_texts))

    return "\n\n".join(sheet_texts).strip()


def _read_docx_as_text(path: Path) -> str:
    """Extract a lightweight text view from a DOCX document.

    This is intentionally simpler and faster than preserving all formatting. It
    extracts paragraph text and skips table content to reduce parsing overhead
    for quick submission/demo runs.
    """

    try:
        import docx
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "python-docx is required to parse .docx files. Install it with `pip install python-docx`."
        ) from exc

    document = docx.Document(path)

    parts: list[str] = []
    for paragraph in document.paragraphs:
        text = paragraph.text.strip()
        if text:
            parts.append(text)

    return "\n\n".join(parts).strip()
